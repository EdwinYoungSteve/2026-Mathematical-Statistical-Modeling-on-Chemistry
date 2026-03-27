import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

INPUT_FILE = r"c:\Users\yxy24\Downloads\DOWNLOAD-wYGnH0T1L6RlO5OLpIrUTS-vFxlmGmQXYyaMVhCnXB8_eq_\DOWNLOAD-wYGnH0T1L6RlO5OLpIrUTS-vFxlmGmQXYyaMVhCnXB8_eq_.csv"
OUTPUT_FILE = r"c:\Users\yxy24\Downloads\EGFR_bioactivity_data.xlsx"

# Read file
with open(INPUT_FILE, "r", encoding="utf-8") as f:
    lines = f.readlines()

# Parse the semicolon-separated data inside the outer quotes
# Each line looks like: "field1;""field2"";""field3"";..."   ,,,,,
def parse_line(line):
    # Remove trailing commas and whitespace
    line = line.rstrip().rstrip(",").strip()
    # Remove outer quotes if present
    if line.startswith('"') and line.endswith('"'):
        line = line[1:-1]
    # Split by ;
    # Fields are separated by ; and each field is wrapped in "" (doubled quotes)
    parts = line.split(";")
    result = []
    for p in parts:
        p = p.strip()
        # Remove surrounding "" pairs
        if p.startswith('""') and p.endswith('""'):
            p = p[2:-2]
        elif p.startswith('"') and p.endswith('"'):
            p = p[1:-1]
        result.append(p)
    return result

# Parse header
header_fields = parse_line(lines[0])

# We need to handle the long assay description that spans "lines" within a single CSV row.
# The data has embedded quotes and commas in the assay description field.
# Strategy: read each logical record by tracking quote balance

def parse_records(lines):
    """Parse logical records from the file, handling multiline quoted fields."""
    records = []
    current = ""
    for line in lines:
        current += line
        # Count unescaped quotes - if balanced, we have a complete record
        # Simple heuristic: a complete record starts with " and the main data ends with a clear pattern
        # Actually, let's just try parsing and see if we get the right number of fields
        stripped = current.rstrip().rstrip(",").strip()
        if stripped.startswith('"'):
            # Check if we have a balanced record
            # The record ends when the quoted section closes
            # Let's count: the line should end with a quote followed by commas
            if re.search(r'""[,\s]*$', current.rstrip()):
                records.append(current)
                current = ""
            elif not stripped.startswith('"') or stripped.endswith('"'):
                records.append(current)
                current = ""
        else:
            records.append(current)
            current = ""
    if current.strip():
        records.append(current)
    return records

all_records = parse_records(lines)

# Parse header
header = parse_line(all_records[0])

# Clean header - remove empty trailing fields
while header and header[-1] == "":
    header.pop()

# Parse data rows
data_rows = []
for record in all_records[1:]:
    fields = parse_line(record)
    # Trim to header length
    fields = fields[:len(header)]
    # Pad if shorter
    while len(fields) < len(header):
        fields.append("")
    data_rows.append(fields)

print(f"Header has {len(header)} columns")
print(f"Found {len(data_rows)} data rows")
print(f"Header: {header[:10]}...")  # Print first 10 columns

# Select key columns for a cleaner output
# Important columns for drug discovery:
key_cols = [
    "Molecule ChEMBL ID",
    "Molecule Name",
    "Molecular Weight",
    "#RO5 Violations",
    "AlogP",
    "Compound Key",
    "Smiles",
    "Standard Type",
    "Standard Relation",
    "Standard Value",
    "Standard Units",
    "pChEMBL Value",
    "Ligand Efficiency BEI",
    "Ligand Efficiency LE",
    "Ligand Efficiency LLE",
    "Ligand Efficiency SEI",
    "Assay ChEMBL ID",
    "Assay Type",
    "BAO Label",
    "Assay Organism",
    "Target ChEMBL ID",
    "Target Name",
    "Target Organism",
    "Target Type",
    "Document ChEMBL ID",
    "Source Description",
    "Document Journal",
    "Document Year",
    "Action Type",
    "Value",
]

# Find column indices
col_indices = []
available_key_cols = []
for col in key_cols:
    if col in header:
        col_indices.append(header.index(col))
        available_key_cols.append(col)

# Create workbook
wb = openpyxl.Workbook()

# ===== Sheet 1: Key Data (clean summary) =====
ws1 = wb.active
ws1.title = "关键数据"

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Write header
for j, col_name in enumerate(available_key_cols, 1):
    cell = ws1.cell(row=1, column=j, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Write data
data_font = Font(name="Calibri", size=10)
data_align = Alignment(vertical="center", wrap_text=False)
num_align = Alignment(horizontal="right", vertical="center")

for i, row in enumerate(data_rows, 2):
    for j, idx in enumerate(col_indices, 1):
        val = row[idx] if idx < len(row) else ""
        # Try to convert numeric values
        try:
            if "." in val:
                val = float(val)
            elif val.isdigit():
                val = int(val)
        except (ValueError, AttributeError):
            pass
        cell = ws1.cell(row=i, column=j, value=val if val != "None" and val != "" else None)
        cell.font = data_font
        cell.border = thin_border
        if isinstance(val, (int, float)):
            cell.alignment = num_align
        else:
            cell.alignment = data_align

# Auto-adjust column widths
for j, col_name in enumerate(available_key_cols, 1):
    max_len = len(str(col_name))
    for i in range(2, len(data_rows) + 2):
        cell_val = ws1.cell(row=i, column=j).value
        if cell_val:
            cell_len = len(str(cell_val))
            if cell_len > max_len:
                max_len = cell_len
    # Cap width, especially for SMILES
    ws1.column_dimensions[openpyxl.utils.get_column_letter(j)].width = min(max_len + 2, 50)

# Freeze top row
ws1.freeze_panes = "A2"
# Auto-filter
ws1.auto_filter.ref = ws1.dimensions

# ===== Sheet 2: All Data =====
ws2 = wb.create_sheet("完整数据")

for j, col_name in enumerate(header, 1):
    cell = ws2.cell(row=1, column=j, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for i, row in enumerate(data_rows, 2):
    for j, val in enumerate(row, 1):
        try:
            if val and "." in val:
                val = float(val)
            elif val and val.isdigit():
                val = int(val)
        except (ValueError, AttributeError):
            pass
        cell = ws2.cell(row=i, column=j, value=val if val != "None" and val != "" else None)
        cell.font = data_font
        cell.border = thin_border

for j in range(1, len(header) + 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width = min(20, 50)

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

# Save
wb.save(OUTPUT_FILE)
print(f"\nExcel file saved to: {OUTPUT_FILE}")
print(f"Sheet 1 '关键数据': {len(available_key_cols)} columns x {len(data_rows)} rows")
print(f"Sheet 2 '完整数据': {len(header)} columns x {len(data_rows)} rows")
